#!/usr/bin/env python3

# external libraries
import sys
import os
import csv
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, numbers

# internal libraries

# colors
orange = PatternFill(start_color='ffcaa1',end_color='ffcaa1',fill_type='solid')
purple = PatternFill(start_color='dbb5ff',end_color='dbb5ff',fill_type='solid')
red1   = PatternFill(start_color='ffa1a1',end_color='ffa1a1',fill_type='solid')
red2   = PatternFill(start_color='e06565',end_color='e06565',fill_type='solid')
red3   = PatternFill(start_color='ff2e2e',end_color='ff2e2e',fill_type='solid')
yellow = PatternFill(start_color='fffa99',end_color='fffa99',fill_type='solid')
green1 = PatternFill(start_color='c1ffab',end_color='c1ffab',fill_type='solid')
green2 = PatternFill(start_color='00ff80',end_color='00ff80',fill_type='solid')
green3 = PatternFill(start_color='34fa4f',end_color='34fa4f',fill_type='solid')
blue   = PatternFill(start_color='abf2ff',end_color='abf2ff',fill_type='solid')
white  = PatternFill(start_color='ffffff',end_color='ffffff',fill_type='solid')
black  = PatternFill(start_color='000000',end_color='000000',fill_type='solid')

# global variables
tran_mon                = input("enter month of transactions: ")
bmo_csv_mc              = ""
bmo_csv_mc_sheet_name   = ""
cibc_csv_chq            = ""
cibc_csv_chq_sheet_name = ""
cibc_csv_visa           = ""
cibc_csv_visa_sheet_name= ""
# template_file_name       = ""
# template_file_sheet_name = ""

output_file             = tran_mon + "-transactions.xlsx"
output_file_sheet_title = tran_mon + "-aggregate"

wb = Workbook()
ws = wb.active
wb.save(output_file)

"""###################################################
store data input by user into global variables
###################################################"""
def process_user_input():
    global tran_mon, bmo_csv_mc, cibc_csv_chq, cibc_csv_visa
    bmo_csv_mc = input("enter BMO transactions file name     : ")
    cibc_csv_chq = input("enter CIBC CHQ transactions file name: ")
    cibc_csv_visa = input("enter CIBC VISA transactions file name: ")

    # global template_file_name
    # template_file_name = input("enter template transactions file name: ")

"""###################################################
process bmo csv file into a consistent format
bmo_mc sheet composition:
- col A: date
- col B: transaction detail
- col C: debit transaction
- col D: credit transaction
- col E: transaction type
###################################################"""
def process_bmo_mc(input_file):
    print("\n1) process BMO transactions file: {}".format(input_file))
    wb = load_workbook(output_file)
    global bmo_csv_mc_sheet_name
    bmo_csv_mc_sheet_name = tran_mon + "-bmo mc"
    ws = wb.create_sheet(bmo_csv_mc_sheet_name)

    # copy from csv to output_file
    with open(input_file,'r') as f:
        for row in csv.reader(f):
            ws.append(row)

    # delete row 1
    ws.delete_rows(1,1)

    # fix transaction date and print to column 1
    i = 1
    while i <= ws.max_row:
        ws.cell(row=i, column=1).value = '=DATE(LEFT(C{},4), MID(C{},5,2), RIGHT(C{},2))'.format(i,i,i)
        i +=1

    # convert transaction date value to data (from formula)
    ws.title = bmo_csv_mc_sheet_name
    wb.save(output_file)
    reply = input("open {}, make col A values only and then press enter ".format(output_file))
    wb = load_workbook(output_file)
    ws = wb[bmo_csv_mc_sheet_name]

    # set format for transaction date
    i = 1
    while i <= ws.max_row:
        ws.cell(row=i, column=1).number_format = numbers.FORMAT_DATE_YYYYMMDD2
        i +=1

    # move transaction description to column 2
    ws.move_range("F1:F{}".format(ws.max_row), rows=0, cols=-4)

    # move transaction amount to column 3 + convert text to number + set format
    ws.move_range("E1:E{}".format(ws.max_row), rows=0, cols=-2)
    i = 1
    while i <= ws.max_row:
        ws.cell(row=i, column=3).value = float(ws.cell(row=i, column=3).value)
        ws.cell(row=i, column=3).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        i +=1


    # if transaction amount -ve, move to column 4
    i = 1
    while i <= ws.max_row:
        transaction = ws.cell(row=i, column=3).value
        if (transaction < 0):
            ws.cell(row=i, column=4).value = (ws.cell(row=i, column=3).value * -1)
            ws.cell(row=i, column=4).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            ws.cell(row=i, column=3).value = 0
            ws.cell(row=i, column=3).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        elif(transaction >= 0):
            ws.cell(row=i, column=4).value = 0
            ws.cell(row=i, column=4).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        i +=1

    # set transaction type in column 5
    i = 1
    while i <= ws.max_row:
        ws.cell(row=i, column=5).value = 'BMO MC'
        i +=1

    # save workbook
    wb.save(output_file)
    print("done")

"""###################################################
process cibc chq csv file into a consistent format
cibc chq sheet composition:
- col A: date
- col B: transaction detail
- col C: debit transaction
- col D: credit transaction
- col E: transaction type
###################################################"""
def process_cibc_chq(input_file):
    print("\n2) process CIBC CHQ transactions file: {}".format(input_file))
    wb = load_workbook(output_file)
    global cibc_csv_chq_sheet_name
    cibc_csv_chq_sheet_name = tran_mon + "-cibc chq"
    ws = wb.create_sheet(cibc_csv_chq_sheet_name)

    # copy from csv to output_file
    with open(input_file,'r') as f:
        for row in csv.reader(f):
            ws.append(row)

    # format column 3 & 4 currency + add transaction type
    i = 1
    while i <= ws.max_row:

        if ws.cell(row=i, column=3).value is "":
            ws.cell(row=i, column=3).value = 0
        
        if ws.cell(row=i, column=4).value is "":
            ws.cell(row=i, column=4).value = 0

        ws.cell(row=i, column=3).value = float(ws.cell(row=i, column=3).value)
        ws.cell(row=i, column=4).value = float(ws.cell(row=i, column=4).value)

        ws.cell(row=i, column=3).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws.cell(row=i, column=4).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws.cell(row=i, column=5).value = 'CIBC CHQ'
        i +=1

    # save workbook
    wb.save(output_file)
    print("done")

"""###################################################
process cibc visa csv file into a consistent format
cibc visa sheet composition:
- col A: date
- col B: transaction detail
- col C: debit transaction
- col D: credit transaction
- col E: transaction type
###################################################"""
def process_cibc_visa(input_file):
    print("\n3) process CIBC VISA transactions file: {}".format(input_file))
    wb = load_workbook(output_file)
    global cibc_csv_visa_sheet_name
    cibc_csv_visa_sheet_name = tran_mon + "-cibc visa"
    ws = wb.create_sheet(cibc_csv_visa_sheet_name)

    # copy from csv to output_file
    with open(input_file,'r') as f:
        for row in csv.reader(f):
            ws.append(row)

    # format column 3 & 4 currency + add transaction type
    i = 1
    while i <= ws.max_row:

        if ws.cell(row=i, column=3).value is "":
            ws.cell(row=i, column=3).value = 0
        
        if ws.cell(row=i, column=4).value is "":
            ws.cell(row=i, column=4).value = 0

        ws.cell(row=i, column=3).value = float(ws.cell(row=i, column=3).value)
        ws.cell(row=i, column=4).value = float(ws.cell(row=i, column=4).value)

        ws.cell(row=i, column=3).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws.cell(row=i, column=4).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        
        ws.cell(row=i, column=5).value = 'CIBC VISA'
        i +=1

    # save workbook
    wb.save(output_file)
    print("done")

"""###################################################
sample template process function
###################################################"""
# def process_template(template_file_name):
    # wb = load_workbook(output_file)
    # global template_file_sheet_name
    # template_file_sheet_name = tran_mon + "-xxxx"
    # ws = wb.create_sheet(template_file_sheet_name)

"""###################################################
###################################################"""
def construct_monthly_transactions():
    print("\n4) construct {} sheet in {}".format(output_file_sheet_title,output_file))
    wb = load_workbook(output_file)
    ws_dst = wb.create_sheet(output_file_sheet_title)
    dst_row_cnt = 0

    # read from bmo csv file
    ws_src1 = wb[bmo_csv_mc_sheet_name]
    maxr = ws_src1.max_row
    maxc = ws_src1.max_column
    for r in range(1, maxr+1):
        for c in range(1, maxc+1):
            ws_dst.cell(row=(r+dst_row_cnt),column=c).value = ws_src1.cell(row=r,column=c).value

    # read from cibc-chq file
    dst_row_cnt = dst_row_cnt + maxr
    ws_src2 = wb[cibc_csv_chq_sheet_name]
    maxr = ws_src2.max_row
    maxc = ws_src2.max_column
    for r in range(1, maxr+1):
        for c in range(1, maxc+1):
            ws_dst.cell(row=(r+dst_row_cnt),column=c).value = ws_src2.cell(row=r,column=c).value

    # read from cibc-visa file
    dst_row_cnt = dst_row_cnt + maxr
    ws_src3 = wb[cibc_csv_visa_sheet_name]
    maxr = ws_src3.max_row
    maxc = ws_src3.max_column
    for r in range(1, maxr+1):
        for c in range(1, maxc+1):
            ws_dst.cell(row=(r+dst_row_cnt),column=c).value = ws_src3.cell(row=r,column=c).value

    # re-format column A, C and D since format may get corrupted in inter-sheet copying
    i = 1
    while i <= ws_dst.max_row:
        ws_dst.cell(row=i, column=1).number_format = numbers.FORMAT_DATE_YYYYMMDD2
        ws_dst.cell(row=i, column=3).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws_dst.cell(row=i, column=4).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE        
        i +=1

    wb.save(output_file)
    print("done")

def main():
    process_user_input()
    process_bmo_mc(bmo_csv_mc)
    process_cibc_chq(cibc_csv_chq)
    process_cibc_visa(cibc_csv_visa)
    construct_monthly_transactions()

if __name__== "__main__":
  main()